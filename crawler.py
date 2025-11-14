import json
import jsonlines
from datetime import datetime
from typing import List, Dict, Optional
import requests
from bs4 import BeautifulSoup
import time
import csv
import re
import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import random


def get_executable_dir():
    """실행 파일이 있는 디렉토리 경로 반환"""
    # 환경 변수 OUTPUT_DIR 확인 (배포 환경용)
    output_dir = os.getenv("OUTPUT_DIR")
    if output_dir:
        return output_dir
    
    if getattr(sys, 'frozen', False):
        # PyInstaller로 실행된 경우
        return os.path.dirname(sys.executable)
    else:
        # 일반 Python으로 실행된 경우
        return os.path.dirname(os.path.abspath(__file__))

class CrawlerCancelledException(Exception):
    """사용자 취소 예외"""
    pass


class PriceCompareCrawler:
    def __init__(
        self,
        config_file: str = None,
        results_file: str = None,
        site_name: str = None,
        cancel_event: Optional[threading.Event] = None
    ):

        self.site_name = site_name
        
        if config_file is None:
            config_file = f"{site_name}_input_list.jsonl"
        self.config_file = config_file

        if results_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")    
            exe_dir = get_executable_dir()

            # JSONL은 임시 폴더에 저장
            temp_dir = tempfile.gettempdir()
            self.results_file = os.path.join(temp_dir, f"{site_name}_가격조사_{timestamp}.jsonl")
            
            # Excel 파일은 OUTPUT_DIR 또는 실행 디렉토리에 저장
            # 배포 환경에서는 OUTPUT_DIR 환경 변수 사용 권장
            self.csv_file = os.path.join(exe_dir, f"{site_name}_가격조사_{timestamp}.xlsx")
        else:
            self.results_file = results_file
            self.csv_file = results_file.replace('.jsonl', '.xlsx')  

        self.progress = 0  # 진행율 저장
        self.total_products = 0  # 전체 제품 수
        self.current_product = 0  # 현재 처리 중인 제품 번호
        self.progress_lock = threading.Lock()  # 진행률 업데이트용 락
        self.cancel_event = cancel_event or threading.Event()
        
    def load_products(self) -> List[Dict]:
        """JSONL 파일에서 제품 정보 로드"""
        products = []
        try:
            with jsonlines.open(self.config_file) as reader:
                for obj in reader:
                    products.append(obj)
        except FileNotFoundError:
            print(f"{self.config_file} 파일이 없습니다.")
        return products
    
    def get_progress(self) -> Dict:
        """현재 진행율 반환"""
        return {
            'current': self.current_product,
            'total': self.total_products,
            'percentage': self.progress
        }
    
    def crawl_ssg(self, url: str) -> Dict:
        """SSG에서 가격 정보 크롤링"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        try:
            # 재시도 로직 추가
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = requests.get(url, headers=headers, timeout=20)
                    response.raise_for_status()  # HTTP 에러 체크
                    break
                except (requests.exceptions.RequestException, requests.exceptions.Timeout) as e:
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)  # 지수 백오프
                        continue
                    else:
                        raise
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # SSG 가격 정보 추출 
            price_elem = soup.select_one('.cdtl_new_price.notranslate .ssg_price')
            if price_elem:
                product_price_text = price_elem.text.replace(',', '').replace('원', '').replace(' ', '').strip()
                try:
                    product_price = int(product_price_text)
                except ValueError:
                    product_price = None
            else:
                product_price = None

            # 배송비 추출
            delivery_elem = soup.select_one('.cdtl_dl.cdtl_delivery_fee')

            if delivery_elem:
                # 첫 번째 li 요소만 선택
                first_li = delivery_elem.select_one('li')
    
                if first_li:
                    delivery_price_elem = first_li.select_one('em.ssg_price')
                    if delivery_price_elem:
                        numbers = re.sub(r'[^\d]', '', delivery_price_elem.text)
                        delivery_price = int(numbers) if numbers else 0
                        delivery_status = "유료"
                    else:
                        delivery_price = 0
                        delivery_status = "무료"
                else:
                    delivery_price = 0
                    delivery_status = "무료"
            else:
                delivery_price = 0
                delivery_status = "정보 없음"

            # 총 가격 계산
            if product_price is not None:
                total_price = product_price + delivery_price
            else:
                total_price = None
            
            return {
                '상품 url': url,
                '상품 가격': product_price,
                '배송비': delivery_price,
                '배송비 여부': delivery_status,
                '최종 가격': total_price,
                '추출 날짜': datetime.now().isoformat()
            }
        except Exception as e:
            return {
                '상품 url': None,
                '상품 가격': None,
                '배송비': None,
                '배송비 여부': None,
                '최종 가격': None,
                '에러 발생': str(e),
                '추출 날짜': datetime.now().isoformat()
            }

    def crawl_ssg_shoping(self, url: str) -> Dict:
        """SSG shoping에서 가격 정보 크롤링"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        try:
            # 재시도 로직 추가
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = requests.get(url, headers=headers, timeout=20)
                    response.raise_for_status()  # HTTP 에러 체크
                    break
                except (requests.exceptions.RequestException, requests.exceptions.Timeout) as e:
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)  # 지수 백오프
                        continue
                    else:
                        raise
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # SSG 가격 정보 추출 
            price_elem = soup.select_one('.price--3')

            if price_elem:
                sale_price_elem = price_elem.select_one('._salePrice')
                best_price_elem = price_elem.select_one('._bestPrice')

                # ✅ 우선순위: salePrice > bestPrice
                if sale_price_elem:
                    product_price_text = sale_price_elem.get_text(strip=True)
                elif best_price_elem:
                    product_price_text = best_price_elem.get_text(strip=True)
                else:
                    product_price_text = None
                
                # 문자열 정제 및 숫자 변환
                if product_price_text:
                    product_price_text = (
                    product_price_text
                    .replace(',', '')
                    .replace('원', '')
                    .replace(' ', '')
                    .strip()
                    )
                    try:
                        product_price = int(product_price_text)
                    except ValueError:
                        product_price = None
                else:
                    product_price = None
            else:
                product_price = None

            total_price = product_price
            delivery_status = "SSG shoping은 배송비가 없습니다"
            
            return {
                '상품 url': url,
                '상품 가격': product_price,
                '배송비': 0,
                '배송비 여부': delivery_status,
                '최종 가격': total_price,
                '추출 날짜': datetime.now().isoformat()
            }
        except Exception as e:
            return {
                '상품 url': None,
                '상품 가격': None,
                '배송비': None,
                '배송비 여부': None,
                '최종 가격': None,
                '에러 발생': str(e),
                '추출 날짜': datetime.now().isoformat()
            }


    def crawl_price(self, url: str) -> Dict:
        """현재 사이트에 맞는 크롤링 함수 호출"""
        if self.site_name == 'ssg':
            return self.crawl_ssg(url)
        elif self.site_name == 'ssg_shoping':
            return self.crawl_ssg_shoping(url)
        elif self.site_name == 'samsung':
            return self.crawl_samsung(url)
    
    def request_cancel(self):
        """취소 요청"""
        self.cancel_event.set()

    def is_cancelled(self) -> bool:
        return self.cancel_event.is_set()

    def _ensure_not_cancelled(self):
        if self.cancel_event.is_set():
            raise CrawlerCancelledException("Crawler cancelled by user")

    def crawl_single_product(self, product: Dict) -> Dict:
        """단일 제품 크롤링 (멀티스레드용)"""
        result = {
            'product_id': product['product_id'],
            'product_name': product['product_name'],
            'timestamp': datetime.now().isoformat(),
            'prices': []
        }
        
        self._ensure_not_cancelled()

        try:
            # Waffle (우리 회사) 크롤링
            if 'waffle' in product:
                self._ensure_not_cancelled()
                waffle_data = self.crawl_price(product['waffle']['url'])
                result['prices'].append({
                    'seller': 'waffle',
                    **waffle_data   
                })
                # 랜덤 딜레이 (1-2초) - Rate Limiting 방지
                time.sleep(random.uniform(1.0, 2.0))
            
            # 경쟁사 크롤링
            if 'competitors' in product:
                for competitor in product['competitors']:
                    self._ensure_not_cancelled()
                    comp_data = self.crawl_price(competitor['url'])
                    result['prices'].append({
                        'seller': competitor['name'],
                        **comp_data
                    })
                    # 랜덤 딜레이 (1-2초) - Rate Limiting 방지
                    time.sleep(random.uniform(1.0, 2.0))
        except Exception as e:
            # 에러 발생 시 로깅
            print(f"  ⚠️ 제품 {product.get('product_name', 'Unknown')} 크롤링 중 오류: {e}")
            result['error'] = str(e)
        
        # 진행률 업데이트 (스레드 안전)
        with self.progress_lock:
            self.current_product += 1
            self.progress = int((self.current_product / self.total_products) * 100)
            print(f"\n[{self.progress}%] 크롤링 완료 ({self.current_product}/{self.total_products}): {product.get('product_name', 'Unknown')}")
        
        return result



    def run_crawling(self, max_workers: int = None) -> bool:
        """전체 제품에 대해 크롤링 실행 (멀티스레드)"""
        # 워커 수 설정 (환경 변수 또는 기본값)
        if max_workers is None:
            max_workers = int(os.getenv("CRAWLER_WORKERS", "7"))  # 기본값 3개 (안정성 우선)
        
        # 워커 수 제한 (1-5개 권장, 안정성을 위해 3개 권장)
        max_workers = max(1, min(max_workers, 7))
        
        products = self.load_products()
        self.total_products = len(products)
        self.current_product = 0
        self.progress = 0
        
        print(f"\n=== 멀티스레드 크롤링 시작 ===")
        print(f"전체 제품 수: {self.total_products}")
        print(f"워커 수: {max_workers}")
        print(f"예상 속도 향상: 약 {max_workers}배")
        print(f"⚠️ 안정성을 위해 요청 간격: 1-2초 (랜덤)\n")
        
        # 결과를 순서대로 저장하기 위한 딕셔너리
        results_dict = {}
        results_lock = threading.Lock()
        error_count = 0
        error_lock = threading.Lock()
        cancelled = False
        
        with jsonlines.open(self.results_file, mode='w') as writer:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # 모든 제품에 대해 Future 제출
                future_to_product = {
                    executor.submit(self.crawl_single_product, product): idx 
                    for idx, product in enumerate(products)
                }
                
                # 완료된 작업부터 처리
                for future in as_completed(future_to_product):
                    idx = future_to_product[future]
                    try:
                        result = future.result()
                        # 에러가 포함된 경우 카운트
                        if 'error' in result:
                            with error_lock:
                                error_count += 1
                        # 결과를 인덱스 순서로 저장
                        with results_lock:
                            results_dict[idx] = result
                    except CrawlerCancelledException:
                        cancelled = True
                        print("사용자 취소 요청을 감지하여 크롤링을 중단합니다.")
                        break
                    except Exception as e:
                        print(f"  ❌ 제품 {idx} 크롤링 중 심각한 오류 발생: {e}")
                        import traceback
                        traceback.print_exc()
                        # 에러가 발생해도 빈 결과 저장
                        with results_lock:
                            results_dict[idx] = {
                                'product_id': products[idx].get('product_id', idx),
                                'product_name': products[idx].get('product_name', 'Unknown'),
                                'timestamp': datetime.now().isoformat(),
                                'prices': [],
                                'error': str(e)
                            }
                        with error_lock:
                            error_count += 1
            
            # 인덱스 순서대로 결과 파일에 저장
            for idx in range(len(products)):
                if idx in results_dict:
                    writer.write(results_dict[idx])
        
        self.progress = 100
        if cancelled or self.is_cancelled():
            print("\n⏹ 크롤링이 사용자 요청으로 취소되었습니다.")
            print(f"현재까지 처리된 제품: {self.current_product}/{self.total_products}")
            return True

        print(f"\n✓ 크롤링 완료! 결과: {self.results_file}")
        print(f"총 처리된 제품: {self.current_product}/{self.total_products}")
        if error_count > 0:
            print(f"⚠️ 에러 발생한 제품: {error_count}개")
            print(f"   성공률: {((self.total_products - error_count) / self.total_products * 100):.1f}%")
        else:
            print(f"✅ 모든 제품 크롤링 성공!")
        return False
    
    def analyze_prices(self):
        """가격 분석 및 비교"""
        results = []
        with jsonlines.open(self.results_file) as reader:
            for obj in reader:
                results.append(obj)
        
        for result in results:
            print(f"\n{'='*50}")
            print(f"제품: {result['product_name']}")
            print(f"{'='*50}")
            
            waffle_price = None
            competitor_prices = []
            
            for p in result['prices']:
                if p['seller'] == 'waffle':
                    waffle_price = p
                    print(f"\n[Waffle - 우리회사]")
                    print(f"  상품 url: {p['상품 url']}")
                    print(f"  상품 가격: {p['상품 가격']}")
                    print(f"  배송비: {p['배송비']}")
                    print(f"  배송비 여부: {p['배송비 여부']}")
                    print(f"  최종 가격: {p['최종 가격']}")
                else:
                    competitor_prices.append(p)
                    print(f"\n[경쟁사 - {p['seller']}]")
                    print(f"  상품 url: {p['상품 url']}")
                    print(f"  상품 가격: {p['상품 가격']}")
                    print(f"  배송비: {p['배송비']}")
                    print(f"  배송비 여부: {p['배송비 여부']}")
                    print(f"  최종 가격: {p['최종 가격']}")
            
            # 간단한 가격 비교
            if waffle_price and competitor_prices:
                print(f"\n📊 가격 비교 분석")
                print(f"   우리 회사와 경쟁사 {len(competitor_prices)}곳의 가격을 비교했습니다.")



    def export_to_excel_format(self, excel_file: str = None):
        """Excel 파일 생성 - Sheet1: 전체 결과, Sheet2: 가격 역전 항목"""
        try:
            if excel_file is None:
                excel_file = self.csv_file.replace('.csv', '.xlsx')
            
            results = []
            with jsonlines.open(self.results_file) as reader:
                for obj in reader:
                    results.append(obj)
            
            if not results:
                print("변환할 데이터가 없습니다.")
                return
            
            # Excel 워크북 생성
            wb = openpyxl.Workbook()
            
            # Sheet1: 전체 결과
            ws1 = wb.active
            ws1.title = "전체 결과"
            
            # Sheet2: 가격 역전 항목
            ws2 = wb.create_sheet("가격 역전 항목")
            
            # Sheet1 작성
            row = 1
            for result in results:
                cell1 = ws1.cell(row, 1, f"제품명: {result['product_name']}")
                cell2 = ws1.cell(row, 2, f"제품ID: {result['product_id']}")
                cell1.font = Font(color="FF0000FF", bold=True)
                cell2.font = Font(color="FF0000FF", bold=True)
                row += 1
                
                ws1.cell(row, 1, f"추출 시간: {result['timestamp']}")
                row += 1
                row += 1  # 빈 줄
                
                # 헤더
                headers = ['판매처', '상품 url', '상품가격', '배송비', '배송비여부', '최종가격']
                for col, header in enumerate(headers, 1):
                    cell = ws1.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                row += 1
                
                # 데이터
                for price_info in result['prices']:
                    seller_name = 'Waffle (우리회사)' if price_info['seller'] == 'waffle' else f"경쟁사 ({price_info['seller']})"
                    ws1.cell(row, 1, seller_name)
                    ws1.cell(row, 2, price_info.get('상품 url', 'N/A'))
                    ws1.cell(row, 3, price_info.get('상품 가격', 'N/A'))
                    ws1.cell(row, 4, price_info.get('배송비', 'N/A'))
                    ws1.cell(row, 5, price_info.get('배송비 여부', 'N/A'))
                    ws1.cell(row, 6, price_info.get('최종 가격', 'N/A'))
                    row += 1
                
                row += 2

            
            # Sheet2 작성: 가격 역전 항목만
            ws2.cell(1, 1, "가격 역전 항목 (경쟁사가 더 저렴한 경우)")
            ws2.cell(1, 1).font = Font(bold=True, size=14)
            row2 = 3
            
            found_cheaper = False
            
            for result in results:
                # Waffle 가격 찾기
                waffle_price = None
                for price_info in result['prices']:
                    if price_info['seller'] == 'waffle':
                        waffle_price = price_info.get('최종 가격')
                        break
                
                if waffle_price is None or not isinstance(waffle_price, (int, float)):
                    continue
                
                # 경쟁사 중 더 저렴한 곳 찾기
                cheaper_competitors = []
                for price_info in result['prices']:
                    if price_info['seller'] != 'waffle':
                        comp_price = price_info.get('최종 가격')
                        if comp_price and isinstance(comp_price, (int, float)) and comp_price < waffle_price:
                            cheaper_competitors.append(price_info)
                
                # 가격 역전이 있는 경우만 Sheet2에 추가
                if cheaper_competitors:
                    found_cheaper = True
                    
                    ws2.cell(row2, 1, f"제품명: {result['product_name']}")
                    ws2.cell(row2, 1).font = Font(bold=True)
                    ws2.cell(row2, 2, f"제품ID: {result['product_id']}")
                    row2 += 1

                    # 헤더
                    headers = ['판매처', '상품 url', '상품가격', '배송비', '배송비여부', '최종가격', '가격차이']
                    for col, header in enumerate(headers, 1):
                        cell = ws2.cell(row2, col, header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    row2 += 1
                    
                    # Waffle 가격 (참고용)
                    ws2.cell(row2, 1, "Waffle (우리회사)")
                    for price_info in result['prices']:
                        if price_info['seller'] == 'waffle':
                            ws2.cell(row2, 2, price_info.get('상품 url', 'N/A'))
                            ws2.cell(row2, 3, price_info.get('상품 가격', 'N/A'))
                            ws2.cell(row2, 4, price_info.get('배송비', 'N/A'))
                            ws2.cell(row2, 5, price_info.get('배송비 여부', 'N/A'))
                            ws2.cell(row2, 6, waffle_price)
                            ws2.cell(row2, 7, "-")
                            break
                    row2 += 1
                    
                    # 더 저렴한 경쟁사들
                    for comp in cheaper_competitors:
                        seller_name = f"경쟁사 ({comp['seller']})"
                        comp_price = comp.get('최종 가격')
                        price_diff = waffle_price - comp_price
                        
                        ws2.cell(row2, 1, seller_name)
                        ws2.cell(row2, 2, comp.get('상품 url', 'N/A'))
                        ws2.cell(row2, 3, comp.get('상품 가격', 'N/A'))
                        ws2.cell(row2, 4, comp.get('배송비', 'N/A'))
                        ws2.cell(row2, 5, comp.get('배송비 여부', 'N/A'))
                        ws2.cell(row2, 6, comp_price)
                        ws2.cell(row2, 7, f"-{price_diff}원 저렴")
                        
                        # 빨간색 강조
                        ws2.cell(row2, 6).font = Font(color="FF0000", bold=True)
                        ws2.cell(row2, 7).font = Font(color="FF0000", bold=True)
                        row2 += 1
                    
                    row2 += 2
            
            if not found_cheaper:
                ws2.cell(row2, 1, "가격 역전 항목이 없습니다. 모든 제품이 경쟁사보다 저렴하거나 동일합니다.")
                ws2.cell(row2, 1).font = Font(color="008000", bold=True)
            
            # 열 너비 자동 조정
            for ws in [ws1, ws2]:
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(excel_file)
            print(f"✓ Excel 파일 생성 완료: {excel_file}")
                # JSONL 파일 삭제
            try:
                if os.path.exists(self.results_file):
                    os.remove(self.results_file)
                    print(f"임시 파일 삭제됨: {self.results_file}")
            except:
                pass  # 삭제 실패해도 무시
            
            return excel_file
            
        except FileNotFoundError:
            print(f"{self.results_file} 파일이 없습니다.")
            return None
        except Exception as e:
            print(f"Excel 변환 중 에러 발생: {e}")
            import traceback
            traceback.print_exc()
            return None 
    
    def get_latest_prices(self, product_id: int) -> Dict:
        """특정 제품의 최신 가격 정보 조회"""
        try:
            with jsonlines.open(self.results_file) as reader:
                for obj in reader:
                    if obj['product_id'] == product_id:
                        return obj
        except FileNotFoundError:
            print(f"{self.results_file} 파일이 없습니다.")
        return None


# 사용 예시
if __name__ == "__main__":
    # 크롤러 실행
    crawler = PriceCompareCrawler()
    
    final_file = f"신세계_가격조사_{datetime.now():%Y%m%d_%H%M%S}"
    crawler.results_file = final_file +".jsonl"

    print("=== 가격 크롤링 시작 ===\n")
    crawler.run_crawling()
    
    print("\n=== 가격 분석 ===")
    crawler.analyze_prices()

    # csv로 변환
    print("\n=== CSV 변환 ===")
    crawler.export_to_excel_format()

    # 특정 제품 조회
    print("\n=== 특정 제품 조회 ===")
    latest = crawler.get_latest_prices(product_id=1)
    if latest:
        print(f"제품: {latest['product_name']}")
        print(f"마지막 업데이트: {latest['timestamp']}")